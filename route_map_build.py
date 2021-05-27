#!/usr/bin/env python3
import openpyxl
import logging
import sys
import os
import re
import ipaddress
import argparse
import json


class Excel:
	# __init__ is the constructor
	def __init__(self):
		self.as_built = "VG_MP_AsBuilt.xlsx"
		self.lisa_file = "DC1_DC2_BGP-RouteControl_v2.xlsx"

	def build_match_rule_1(self):
		os.system("rm ROUTE_MAP_CSV/1-MR.csv")
		f = open("ROUTE_MAP_CSV/1-MR.csv", "a")
		f.write("TENANT" + "\n")
		f.write("TNT_SWP_SDE" + "\n")
		f.write("TNT_SWP_SOE" + "\n")
		f.write("TNT_SWP_GIS" + "\n")
		f.close()

	def build_csv_2_7(self):
		worksheets = []
		os.system("rm ROUTE_MAP_CSV/2-7.csv")
		f = open("ROUTE_MAP_CSV/2-7.csv", "a")
		f.write("TENANT,VRF,ASN1,ASN2,ASN3" + "\n")
		wb = openpyxl.load_workbook(self.lisa_file, data_only=True)

		for sheet in wb:
			worksheets.append(sheet.title)
		wb.close()

		wb.active = worksheets.index("Set Rules")
		ws = wb.active

		row_start = ws.min_row
		row_end = ws.max_row

		vrf_to_asn = {}

		for x in range(row_start + 1, row_end + 1):
			cell_type = 'C' + str(x)
			cell_type_val = ws[cell_type].value

			if cell_type_val is None:
				continue

			cell_rule = 'B' + str(x)
			cell_rule_val = ws[cell_rule].value

			cell_tenant = 'A' + str(x)
			tenant = ws[cell_tenant].value

			cell_asn = 'G' + str(x)
			asn = ws[cell_asn].value

			if cell_type_val == 'as-path':
				a_cell_rule_val = cell_rule_val.split("_")
				vrf = a_cell_rule_val[1]
				if tenant not in vrf_to_asn:
					vrf_to_asn[tenant] = {}
					vrf_to_asn[tenant][vrf] = asn
					f.write(tenant + "," + vrf + "," + str(asn) + "," + str(asn) + "," + str(asn) + '\n')
				else:
					if vrf not in vrf_to_asn[tenant]:
						vrf_to_asn[tenant][vrf] = asn
						f.write(tenant + "," + vrf + "," + str(asn) + "," + str(asn) + "," + str(asn) + '\n')

		f.close()

	def load_l3o_path_from_asbuilt(self):

		worksheets = []
		wb = openpyxl.load_workbook(self.as_built, data_only=True)

		for sheet in wb:
			worksheets.append(sheet.title)
		wb.close()

		wb.active = worksheets.index("l3out_int_profile")
		ws = wb.active

		row_start = ws.min_row
		row_end = ws.max_row

		l3out_mapping = {}

		for x in range(row_start + 1, row_end + 1):
			lip_cell = 'A' + str(x)
			lip = ws[lip_cell].value

			lnp_cell = 'B' + str(x)
			lnp = ws[lnp_cell].value

			l3o_cell = 'C' + str(x)
			l3o = ws[l3o_cell].value

			tenant_cell = 'D' + str(x)
			tenant = ws[tenant_cell].value

			path_cell = 'I' + str(x)
			path = ws[path_cell].value

			peer_cell = 'X' + str(x)
			peer = ws[peer_cell].value

			if tenant not in l3out_mapping:
				l3out_mapping[tenant] = {}
			if l3o not in l3out_mapping[tenant]:
				l3out_mapping[tenant][l3o] = {}
			if lnp not in l3out_mapping[tenant][l3o]:
				l3out_mapping[tenant][l3o][lnp] = {}
			if lip not in l3out_mapping[tenant][l3o][lnp]:
				l3out_mapping[tenant][l3o][lnp][lip] = {}
			if peer not in l3out_mapping[tenant][l3o][lnp][lip]:
				l3out_mapping[tenant][l3o][lnp][lip][peer] = path


		return l3out_mapping

	def build_csv_8(self, as_built_paths):
		worksheets = []
		os.system("rm ROUTE_MAP_CSV/8-RM.csv")
		f = open("ROUTE_MAP_CSV/8-RM.csv", "a")
		f.write("TENANT,VRF,L3OUT,LNP,LIP,PATH,RM,DIRECTION" + "\n")
		wb = openpyxl.load_workbook(self.lisa_file, data_only=True)

		for sheet in wb:
			worksheets.append(sheet.title)
		wb.close()

		wb.active = worksheets.index("BGP Connectivity Profiles")
		ws = wb.active

		row_start = ws.min_row
		row_end = ws.max_row

		for x in range(row_start + 1, row_end + 1):
			cell_tenant = 'A' + str(x)
			tenant = ws[cell_tenant].value

			if tenant is None:
				continue

			cell_lnp = 'C' + str(x)
			lnp = ws[cell_lnp].value

			cell_lip = 'D' + str(x)
			lip = ws[cell_lip].value

			cell_peer = 'E' + str(x)
			peer = ws[cell_peer].value

			cell_direction = 'F' + str(x)
			direction = ws[cell_direction].value

			cell_rm = 'G' + str(x)
			rm = ws[cell_rm].value

			l3o = lnp
			l3o = l3o.replace("LNP_DC1_","")
			l3o = l3o.replace("LNP_DC2_","")

			a_vrf = lip
			x_vrf = a_vrf.split("_")
			vrf = x_vrf[-2]


			path = "[" + as_built_paths[tenant][l3o][lnp][lip][peer] + "]/peerP-[" + peer + "]"
			f.write(tenant + "," + vrf + "," + l3o + "," + lnp + "," + lip + "," + path + "," + rm + "," + direction + "\n")

		f.close()

	def build_enable_loopback_9(self, as_built_paths):
		worksheets = []
		lnp_map = {}
		os.system("rm ROUTE_MAP_CSV/9.csv")
		f = open("ROUTE_MAP_CSV/9.csv", "a")
		f.write("TENANT,L3OUT,LNP,PODID,NODEID" + "\n")
		wb = openpyxl.load_workbook(self.lisa_file, data_only=True)

		for sheet in wb:
			worksheets.append(sheet.title)
		wb.close()

		wb.active = worksheets.index("BGP Connectivity Profiles")
		ws = wb.active

		row_start = ws.min_row
		row_end = ws.max_row

		for x in range(row_start + 1, row_end + 1):
			cell_tenant = 'A' + str(x)
			tenant = ws[cell_tenant].value

			if tenant is None:
				continue

			cell_lnp = 'C' + str(x)
			lnp = ws[cell_lnp].value

			cell_lip = 'D' + str(x)
			lip = ws[cell_lip].value

			cell_peer = 'E' + str(x)
			peer = ws[cell_peer].value

			l3o = lnp
			l3o = l3o.replace("LNP_DC1_","")
			l3o = l3o.replace("LNP_DC2_","")

			path = as_built_paths[tenant][l3o][lnp][lip][peer]
			m = re.search('pod-(\d)/', path)
			podid = m.group(1)

			m = re.search('paths-(\d+)/', path)
			nodeid = m.group(1)

			if lnp in lnp_map:
				if podid in lnp_map[lnp]:
					if nodeid in lnp_map[lnp][podid]:
						continue
					else:
						lnp_map[lnp][podid][nodeid] = {}
						f.write(tenant + "," + l3o + "," + lnp + "," + podid + "," + nodeid + "\n")
				else:
					lnp_map[lnp][podid] = {}
			else:
				lnp_map[lnp] = {}


		f.close()


def main():
	data = Excel()
	data.build_match_rule_1()
	data.build_csv_2_7()
	as_built_paths = data.load_l3o_path_from_asbuilt()
	data.build_csv_8(as_built_paths)
	data.build_enable_loopback_9(as_built_paths)

if __name__ == '__main__':
    main()