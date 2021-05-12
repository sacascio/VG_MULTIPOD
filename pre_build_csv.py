#!/usr/bin/env python3
import openpyxl
import logging
import sys
import os
import re
import ipaddress
import argparse


class Excel:
	# __init__ is the constructor
	def __init__(self,filename):
		self.tnt_soe = "TNT_SWP_SOE"
		self.tnt_gis = "TNT_SWP_GIS"
		self.tnt_sde = "TNT_SWP_SDE"
		self.vmmd = "VMM_COMMON-VXRAIL-DC-DVSwitch01"
		self.vccontroller = "swpdcxzpvcs001"
		self.node = "N1"
		self.subject = "Permit_Any"
		self.filename = filename
		self.tab = 'DCX PBR'
		self.slapolicy = 'SLA_POLICY_ICM_5'
		self.list_of_non_symm_vrf = []
		self.list_of_non_symm_vrf.append("SOE_INF")
		self.list_of_non_symm_vrf.append("SOE_B2B")
		self.list_of_non_symm_vrf.append("GIS_MMP")
		self.list_of_non_symm_vrf.append("SDE_SDE")
		self.list_of_non_symm_vrf.append("SDE_PTD")
		self.list_of_non_symm_vrf.append("SDE_LDZ")
		self.list_of_non_symm_vrf.append("SDE_D2B")

		self.list_of_type_b_vrf = []
		self.list_of_type_b_vrf.append("SOE_INF")
		self.list_of_type_b_vrf.append("SOE_B2B")
		self.list_of_type_b_vrf.append("SDE_LDZ")
		self.list_of_type_b_vrf.append("SDE_D2B")

		self.vrf_to_rhg = {}

		self.bd_to_ip = {}

		self.bd_to_ip["SOE_VXR"] = "10.178.79.1"
		self.bd_to_ip["SOE_COM"] = "10.178.79.17"
		self.bd_to_ip["SOE_ITC"] = "10.178.79.33"
		self.bd_to_ip["SOE_BSC"] = "10.178.79.49"
		self.bd_to_ip["SOE_PTM"] = "10.178.79.65"
		self.bd_to_ip["SOE_ACC"] = "10.178.79.81"
		self.bd_to_ip["SOE_DAT"] = "10.178.79.97"
		self.bd_to_ip["SOE_DST"] = "10.178.79.113"
		self.bd_to_ip["SOE_INF"] = "10.178.79.129"
		self.bd_to_ip["SOE_B2B"] = "10.178.79.145"

		self.bd_to_ip["SDE_SDE"] = "10.178.95.1"
		self.bd_to_ip["SDE_LDZ"] = "10.178.95.17"
		self.bd_to_ip["SDE_PTD"] = "10.178.95.33"
		self.bd_to_ip["SDE_D2B"] = "10.178.95.49"

		self.bd_to_ip["GIS_COM"] = "10.178.111.1"
		self.bd_to_ip["GIS_BSC"] = "10.178.111.17"
		self.bd_to_ip["GIS_ENT"] = "10.178.111.33"
		self.bd_to_ip["GIS_PTM"] = "10.178.111.49"
		self.bd_to_ip["GIS_DST"] = "10.178.111.65"
		self.bd_to_ip["GIS_MMP"] = "10.178.111.81"

		if not os.path.isfile(self.filename):
			logging.error(" File %s not found" % self.filename)
			sys.exit(9)

		logging.basicConfig(format='%(levelname)s:%(message)s', level=logging.DEBUG)

		# Remove all csv files
		os.system("rm *.csv")

		# Build headers for the CSVs
		f = open("6a.csv", "a")
		f.write("TENANT,FW,VMMDOMAIN" + "\n")
		f.close()

		f = open("6b.csv", "a")
		f.write("TENANT,FW,VM,VMInterface,VMAdapterName,DEVICE,VCENTER_CONTROLLER" + "\n")
		f.close()

		f = open("6c-Symm.csv", "a")
		f.write("TENANT,FW,VMInterface1,VMInterface2,VMInterface3,VMInterface4,CLUSTERNAME,CLUSTERDEVICE1,CLUSTERDEVICE2,CLUSTERDEVICE3,CLUSTERDEVICE4" + "\n")
		f.close()

		f = open("6c-NoSymm.csv", "a")
		f.write("TENANT,FW,VMInterface1,VMInterface2,CLUSTERNAME,CLUSTERDEVICE1,CLUSTERDEVICE2" + "\n")
		f.close()

		f = open("7.csv", "a")
		f.write("TENANT,SGNAME,FW" + "\n")
		f.close()

		f = open("8.csv", "a")
		f.write("TENANT,HGNAME" + "\n")
		f.close()

		f = open("10-Symm.csv", "a")
		f.write("TENANT,PBRNAME,REDIRECTIP1,REDIRECTMAC1,REDIRECTIP2,REDIRECTMAC2,HEALTHGRP1,HEALTHGRP2,SLAPOLICY" + "\n")
		f.close()

		f = open("10-NoSymm.csv", "a")
		f.write("TENANT,PBRNAME,REDIRECTIP1,REDIRECTMAC1" + "\n")
		f.close()

		f = open("12.csv", "a")
		f.write("TENANT,CONTRACT_NAME,SGNAME,NODENAME,FW,BD,PBR_POLICY,CLUSTER" + "\n")
		f.close()

		f = open("13.csv", "a")
		f.write("TENANT,CONTRACT_NAME,SGNAME,SUBJECT" + "\n")
		f.close()


	def build_virtual_device_6a(self,short_tenant,vrf,vmmd):
		f = open("6a.csv", "a")
		f.write("TNT_SWP_" + short_tenant + "," + "SVD_" + short_tenant + "_PFW_" + vrf + "," + vmmd + "\n")
		f.close()

	def build_virtual_interfaces_6b(self,short_tenant,vrf,vfw,fwinterface,vnic,vccontroller):
		f = open("6b.csv", "a")
		f.write("TNT_SWP_" + short_tenant + "," + "SVD_" + short_tenant + "_PFW_" + vrf + "," + vfw + "," +  "CDI_" + vfw + "_" + fwinterface + "," + "Network adapter " + vnic + "," "CDV_" + short_tenant + "_L4L7_" + vfw + "," + vccontroller + "\n")
		f.close()

	def build_cluster_interfaces_6cSymm(self,short_tenant,vrf,vrf_to_fw):
		f = open("6c-Symm.csv", "a")
		f.write("TNT_SWP_" + short_tenant + "," + "SVD_" + short_tenant + "_PFW_" + vrf + "," )

		# Print interfaces 1-4
		for x in range (0,4):
			f.write("CDI_" + vrf_to_fw[short_tenant + "_" + vrf][x][0] + "_" + vrf_to_fw[short_tenant + "_" + vrf][x][2]  + ",")

		f.write("CLS_" + short_tenant + "_PFW_" + vrf )

		# Print cluster devices
		for x in range (0,4):
			f.write(",CDV_" + short_tenant + "_L4L7_" + vrf_to_fw[short_tenant + "_" + vrf][x][0])

		f.write("\n")
		f.close()

	def build_cluster_interfaces_6cNoSymm(self,short_tenant,vrf,vrf_to_fw):
		f = open("6c-NoSymm.csv", "a")
		f.write("TNT_SWP_" + short_tenant + "," + "SVD_" + short_tenant + "_PFW_" + vrf + "," )

		# Print interfaces 1-2
		for x in range (0,2):
			f.write("CDI_" + vrf_to_fw[short_tenant + "_" + vrf][x][0] + "_" + vrf_to_fw[short_tenant + "_" + vrf][x][2] + ",")

		f.write("CLS_" + short_tenant + "_PFW_" + vrf)

		# Print cluster devices
		for x in range (0,2):
			f.write(",CDV_" + short_tenant + "_L4L7_" + vrf_to_fw[short_tenant + "_" + vrf][x][0])

		f.write("\n")
		f.close()

	def build_sgtemplate_7(self,short_tenant,vrf):
		f = open("7.csv", "a")
		f.write("TNT_SWP_" + short_tenant + "," + "SGT_" + short_tenant + "_PBR_" + vrf + "," + "SVD_" + short_tenant + "_PFW_" + vrf + "\n")
		f.close()

	def build_healthgroup_8(self,short_tenant,vrf,vfw):
		f = open("8.csv", "a")
		f.write("TNT_SWP_" + short_tenant + "," + "RHG_" + vfw + "_" + vrf + "\n")
		f.close()

	def build_pbr_policy_10Symm(self,short_tenant,vrf,fwlist):
		f = open("10-Symm.csv", "a")
		short_fw = fwlist[0][:-3]

		try:
			self.bd_to_ip[short_tenant + "_" + vrf]
		except NameError:
			logging.error(" COULD NOT FIND SUBNET ADDRESS FOR TENANT %s, VRF %s" % (short_tenant,vrf))
			sys.exit(9)

		ip = ipaddress.ip_address(self.bd_to_ip[short_tenant + "_" + vrf])
		ip1 = ip + 3
		ip2 = ip + 6

		f.write("TNT_SWP_" + short_tenant + "," + "PBR_" + short_tenant + "_" + "L4L7_" + short_fw + "_" + vrf + "," + str(ip1) + "," + "00:00:00:11:11:11" + "," + str(ip2) + "," + "00:00:0:33:33:33" + "," + "RHG_" + fwlist[0] + "_" + vrf + "," + "RHG_" + fwlist[1] + "_" + vrf + "," + self.slapolicy + "\n")
		f.close()

	def build_pbr_policy_10NoSymm(self,short_tenant,vrf,fwlist):
		f = open("10-NoSymm.csv", "a")
		short_fw = fwlist[0][:-3]

		try:
			self.bd_to_ip[short_tenant + "_" + vrf]
		except NameError:
			logging.error(" COULD NOT FIND SUBNET ADDRESS FOR TENANT %s, VRF %s" % (short_tenant,vrf))
			sys.exit(9)

		ip = ipaddress.ip_address(self.bd_to_ip[short_tenant + "_" + vrf])
		ip1 = ip + 3

		f.write("TNT_SWP_" + short_tenant + "," + "PBR_" + short_tenant + "_" + "L4L7_" + short_fw + "_" + vrf + "," + str(ip1) + "," + "00:00:00:11:11:11" + "\n")

		f.close()

	def build_device_selection_policy_12(self,short_tenant,vrf,vfw):
		f = open("12.csv","a")
		f.write("TNT_SWP_" + short_tenant + "," + "SGC_SWP_" + short_tenant + "_PBR_" + vrf + "," + "SGT_" + short_tenant + "_PBR_" + vrf + "," + self.node + "," + "SVD_" + short_tenant + "_PFW_" + vrf + "," + "SBD_SWP_" + short_tenant + "_PBR_" + vrf + "," + "PBR_" + short_tenant + "_L4L7_" + vfw + "_" + vrf + "," + "CLS_" + short_tenant + "_PFW_" + vrf + "\n")
		f.close()

	def assign_sg_to_contract_13(self,short_tenant,vrf):
		f = open("13.csv", "a")
		f.write("TNT_SWP_" + short_tenant + "," + "SGC_SWP_" + short_tenant + "_PBR_" + vrf + "," + "SGT_" + short_tenant + "_PBR_" + vrf + "," +  self.subject + "\n")
		f.close()


	def parse_excel(self):
		worksheets = []

		wb = openpyxl.load_workbook(self.filename, data_only=True)

		for sheet in wb:
			worksheets.append(sheet.title)
		wb.close()

		if self.tab not in worksheets:
			logging.error(" Could not find sheet %s in file %s" % (self.tab,self.filename))

		wb.active = worksheets.index(self.tab)
		ws = wb.active

		row_start = ws.min_row
		row_end = ws.max_row

		vrf_to_fw = {}

		# Part 1 - get the VRF to FW mapping - build CSVs from the vrf_to_fw structure
		for x in range(row_start + 1, row_end + 1):
			cell = 'B' + str(x)
			vfw = ws[cell].value

			if vfw is None:
				continue

			if bool(re.search('PPFW', vfw, re.IGNORECASE)):
				tenant_cell = 'I' + str(x)
				long_tenant = ws[tenant_cell].value
				t_x = long_tenant.split("_")
				short_tenant = t_x[2]

				vrf_cell = 'J' + str(x)
				vrf = ws[vrf_cell].value
				v_x = vrf.split("_")
				vrf = v_x[3]

				vnic_cell = 'O' + str(x)
				vnic = str(ws[vnic_cell].value)

				fwinterface_cell = 'F' + str(x)
				fwinterface = ws[fwinterface_cell].value
				fwinterface = fwinterface.replace("/", "_")

				if 	short_tenant + "_" + vrf not in vrf_to_fw:
					vrf_to_fw[short_tenant + "_" + vrf] = []

				if vfw not in vrf_to_fw[short_tenant + "_" + vrf]:
					vrf_to_fw[short_tenant + "_" + vrf].append([vfw,vnic,fwinterface])

		for tenant_vrf in vrf_to_fw:
			tvdata = tenant_vrf.split("_")
			short_tenant = tvdata[0]
			vrf = tvdata[1]

			# 6A - Create device
			self.build_virtual_device_6a(short_tenant, vrf, self.vmmd)

			# 6C - create cluster interfaces - depending if symmetric or non symmetric PBR being used
			if short_tenant + "_" + vrf in self.list_of_non_symm_vrf:
				self.build_cluster_interfaces_6cNoSymm(short_tenant, vrf, vfw, vrf_to_fw)
				logging.info("%s_%s building concrete devices for non symmetric PBR" % (short_tenant, vrf))
			else:
				self.build_cluster_interfaces_6cSymm(short_tenant, vrf, vrf_to_fw)
				logging.info("%s_%s building concrete devices for symmetric PBR" % (short_tenant, vrf))

			# 7 - Build SGT
			self.build_sgtemplate_7(short_tenant, vrf)

			# 8 - Build only 'A' side RHG's

			for vfw in vrf_to_fw[tenant_vrf]:
				if vfw[0].endswith('A'):
					self.build_healthgroup_8(short_tenant, vrf, vfw[0])

					if short_tenant + "_" + vrf not in self.vrf_to_rhg:
						self.vrf_to_rhg[short_tenant + "_" + vrf] = []

					self.vrf_to_rhg[short_tenant + "_" + vrf].append(vfw[0])

			# 10 - PBR Policy depending if symmetric or non symmetric PBR being used
			if short_tenant + "_" + vrf in self.list_of_non_symm_vrf:
				self.build_pbr_policy_10NoSymm(short_tenant,vrf,self.vrf_to_rhg[short_tenant + "_" + vrf])
			else:
				self.build_pbr_policy_10Symm(short_tenant,vrf,self.vrf_to_rhg[short_tenant + "_" + vrf])

			# 12 - Build device selection policy
			self.build_device_selection_policy_12(short_tenant, vrf, vrf_to_fw[tenant_vrf][0][0][:-3])

			# 13 - Assign SGT to contract
			self.assign_sg_to_contract_13(short_tenant,vrf)

		# part 2 -  build CSVs that require line by line analysis

		for x in range(row_start + 1, row_end + 1):
			cell = 'B' + str(x)
			vfw = ws[cell].value

			if vfw is None:
				continue

			if bool(re.search('PPFW', vfw, re.IGNORECASE)):
				tenant_cell = 'I' + str(x)
				long_tenant = ws[tenant_cell].value
				t_x = long_tenant.split("_")
				short_tenant = t_x[2]

				vrf_cell = 'J' + str(x)
				vrf = ws[vrf_cell].value
				v_x = vrf.split("_")
				vrf = v_x[3]

				vnic_cell = 'O' + str(x)
				vnic = str(ws[vnic_cell].value)

				fwinterface_cell = 'F' + str(x)
				fwinterface = ws[fwinterface_cell].value
				fwinterface = fwinterface.replace("/","_")

				# 6B - build interfaces
				self.build_virtual_interfaces_6b(short_tenant, vrf, vfw, fwinterface,vnic, self.vccontroller)


def read_arguments():
	parser = argparse.ArgumentParser("Usage: ./pre_build_csv.py -f <LCS File>")
	parser.add_argument("-f", "--input-file", dest="filename" , help="LCS xlsx file", required=True)
	args = parser.parse_args()
	return args

def main():
	args = read_arguments()
	data = Excel(args.filename)
	data.parse_excel()
if __name__ == '__main__':
    main()